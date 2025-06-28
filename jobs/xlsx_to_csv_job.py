import boto3
import os
from datetime import datetime
import openpyxl

# S3 Configuration
S3_BUCKET = "ecommerce-lakehouse-001"
RAW_PREFIX = "raw/"
OUTPUT_PREFIX = "preprocessed/"
LOCAL_TMP_PATH = "/tmp"

# Initialize S3 client
s3 = boto3.client("s3")

def download_xlsx_from_s3(s3_key, local_path):
    s3.download_file(S3_BUCKET, s3_key, local_path)
    print(f"Downloaded {s3_key} to {local_path}")

def convert_excel_sheets_to_csv(local_xlsx_path, output_prefix):
    """
    Converts each sheet in an Excel file to a separate CSV file and uploads to S3.
    
    Args:
        local_xlsx_path (str): Local path to the downloaded Excel file.
        output_prefix (str): S3 prefix where the CSV files will be uploaded.
    
    Note: Uses openpyxl to read Excel sheets and saves data as CSV with comma separation.
          Handles None values by converting them to empty strings.
    """
    wb = openpyxl.load_workbook(local_xlsx_path, data_only=True)  # Loads workbook, data_only=True to get calculated values.
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        csv_lines = []
        for row in sheet.iter_rows(values_only=True):  # Iterates over rows, getting values only.
            row_str = ",".join(['' if cell is None else str(cell) for cell in row])  # Converts row to CSV line.
            csv_lines.append(row_str)

        csv_filename = f"{sheet_name}.csv"  # Names the CSV file after the sheet.
        local_csv_path = os.path.join(LOCAL_TMP_PATH, csv_filename)  # Local path for the CSV.

        with open(local_csv_path, "w", encoding="utf-8") as f:  # Writes CSV content.
            f.write("\n".join(csv_lines))

        s3_key = f"{output_prefix}{csv_filename}"  # S3 key for the uploaded CSV.
        s3.upload_file(local_csv_path, S3_BUCKET, s3_key)  # Uploads the CSV to S3.
        print(f"📤 Uploaded {csv_filename} to s3://{S3_BUCKET}/{s3_key}")
